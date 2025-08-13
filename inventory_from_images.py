
"""
inventory_from_images.py
------------------------
Reads image files from a folder and generates an Excel file that includes:
  - A thumbnail preview for each image
  - A temporary item name derived from the file name

Usage:
  1) Install dependencies (see README.md or requirements.txt)
  2) Put your images in a folder (default: ./images)
  3) Run:
        python inventory_from_images.py --input ./images --output ./inventory_audit.xlsx
     (Both args are optional; sensible defaults are provided.)
"""

import argparse
import io
import sys
import time
from datetime import datetime
from pathlib import Path

SUPPORTED_EXTS = {".jpg", ".jpeg", ".png", ".webp"}

def main():
    parser = argparse.ArgumentParser(description="Create an inventory Excel from images in a folder.")
    parser.add_argument(
        "--input", "-i", type=str, default="images",
        help="Path to the folder containing images (default: ./images)"
    )
    parser.add_argument(
        "--output", "-o", type=str, default="inventory_audit.xlsx",
        help="Path to the output Excel file (default: ./inventory_audit.xlsx)"
    )
    parser.add_argument(
        "--thumb-size", type=int, default=100,
        help="Thumbnail size in pixels for both width and height (default: 100)"
    )
    parser.add_argument(
        "--timestamped", action="store_true",
        help="Append a YYYYMMDD_HHMMSS timestamp to the output filename (before extension)"
    )
    parser.add_argument(
        "--recursive", action="store_true",
        help="Recursively include images from subfolders"
    )
    args = parser.parse_args()

    try:
        from PIL import Image as PILImage  # Pillow
    except Exception as e:
        print("ERROR: Pillow is not installed. Install with: pip install Pillow", file=sys.stderr)
        raise

    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
    except Exception as e:
        print("ERROR: openpyxl is not installed. Install with: pip install openpyxl", file=sys.stderr)
        raise

    input_path = Path(args.input)
    if not input_path.exists() or not input_path.is_dir():
        print(f"ERROR: Input folder not found or not a directory: {input_path}", file=sys.stderr)
        sys.exit(1)

    # Collect image files
    if args.recursive:
        files = sorted([p for p in input_path.rglob("*") if p.suffix.lower() in SUPPORTED_EXTS])
    else:
        files = sorted([p for p in input_path.iterdir() if p.is_file() and p.suffix.lower() in SUPPORTED_EXTS])

    if not files:
        print(f"WARNING: No supported image files found in: {input_path}", file=sys.stderr)

    # Prepare workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    # Headers
    ws["A1"] = "Photo"
    ws["B1"] = "Temporary Name"
    ws.freeze_panes = "A2"

    # Set column widths (approx); row heights set per-image
    ws.column_dimensions["A"].width = 18  # image column
    ws.column_dimensions["B"].width = 40  # name column

    # Thumbnail temp cache in-memory (BytesIO) to avoid filesystem writes
    # openpyxl accepts PIL images directly, but to be robust across versions,
    # we convert to PNG bytes and pass a BytesIO-backed image where needed.
    row = 2
    for fp in files:
        try:
            with PILImage.open(fp) as im:
                im = im.convert("RGBA")  # ensure consistent mode
                # Make thumbnail (preserve aspect ratio, fit into square box)
                im.thumbnail((args.thumb_size, args.thumb_size))

                # Save to PNG bytes buffer for Excel embedding
                buf = io.BytesIO()
                im.save(buf, format="PNG")
                buf.seek(0)

                # Create openpyxl image from buffer
                xl_img = XLImage(buf)
                # Slight padding in cell
                ws.row_dimensions[row].height = max(80, args.thumb_size * 0.75)  # Excel row height is in points
                # Add image to cell A{row}
                cell_addr = f"A{row}"
                ws.add_image(xl_img, cell_addr)

                # Temporary name from filename (no extension)
                ws[f"B{row}"] = fp.stem
                row += 1
        except Exception as e:
            # Log error but continue
            ws[f"B{row}"] = f"[ERROR reading image] {fp.name}"
            row += 1
            continue

    # Output filename (optionally timestamped)
    out_path = Path(args.output)
    if args.timestamped:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = out_path.with_name(out_path.stem + f"_{ts}" + out_path.suffix)

    # Ensure parent exists
    if out_path.parent and not out_path.parent.exists():
        out_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(out_path)
    print(f"Excel created: {out_path.resolve()}")
    print(f"Images processed: {len(files)}")

if __name__ == "__main__":
    main()
